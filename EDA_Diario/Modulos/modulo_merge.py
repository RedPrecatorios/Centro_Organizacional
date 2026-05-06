"""
Enriquecimento em duas etapas — mesmo fluxo para prc_tjsp, prc_cmp e prc_imp:

  Etapa 1 — `etapa1_enriquecer_com_p2`: principal (via `modelo` em
  `processar_planilha_principal`) + Lemitti (P2) -> INTERMEDIARIA.xlsx e
  cpfs_nao_encontrados_p2.csv (CPFs sem contato na P2).

  Etapa 2 — `etapa2_enriquecer_com_p3`: retoma a intermediária + P3 (ex. Assertiva)
  para enriquecer só quem ainda faltou.

  O `modelo` só afecta a leitura da principal. CMP e IMP usam 2.ª coluna CPF
  no cruzamento; TJSP tipicamente só `CPF`.
"""
import io
import os
import re
import unicodedata
import pandas as pd


def _pular_cooldown_etapa2() -> bool:
    """
    Cooldown desligado se:
      - EDA_SKIP_COOLDOWN=1|true|yes|on|sim
      - EDA_COOLDOWN_DIAS=0 (ou negativo)
    """
    v = (os.environ.get("EDA_SKIP_COOLDOWN") or "").strip().lower()
    if v in ("1", "true", "yes", "on", "sim"):
        return True
    try:
        d = int((os.environ.get("EDA_COOLDOWN_DIAS") or "14").strip())
    except ValueError:
        d = 14
    return d <= 0


def _dias_cooldown_etapa2() -> int:
    """Janela em dias (padrao 14). Com EDA_COOLDOWN_DIAS=0 o filtro e ignorado via _pular_cooldown_etapa2."""
    try:
        d = int((os.environ.get("EDA_COOLDOWN_DIAS") or "14").strip())
    except ValueError:
        d = 14
    return max(1, d) if d > 0 else 14


def _cooldown_permitir_planilha_totalmente_filtrada() -> bool:
    """
    Se true, remove mesmo quando *todos* os registros estao em cooldown (planilha vazia).
    Por defeito e *false*: evita FINAL vazio em reprocessamentos dentro da janela.
    Env: EDA_COOLDOWN_PERMITIR_PLANILHA_VAZIA=1|true|yes|on|sim
    """
    z = (os.environ.get("EDA_COOLDOWN_PERMITIR_PLANILHA_VAZIA") or "").strip().lower()
    return z in ("1", "true", "yes", "on", "sim")
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from modulo_planilha_principal import processar_planilha_principal
from modulo_enriquecimento_contatos import (
    COLUNA_CPF as P2_COL_CPF,
    COLUNA_NOME as P2_COL_NOME,
    ler_serie_telefone_concat_colunas_excel,
    processar_enriquecimento_contatos,
)
from modulo_enriquecimento_relacionados import processar_enriquecimento_relacionados
from modulo_banco import (
    criar_banco_e_tabelas, registrar_execucao, salvar_processos,
    salvar_contatos, salvar_disparo_hsm, carregar_blacklist,
    buscar_cpfs_cooldown,
)
from modulo_blacklist import filtrar_hsm_por_blacklist, filtrar_registros_por_blacklist

# ─────────────────────────────────────────────
# Prefixos das colunas geradas na planilha final
# ─────────────────────────────────────────────
PREFIXO_TELEFONE = "TELEFONE"
PREFIXO_EMAIL    = "EMAIL"
PREFIXO_HSM_LEMITTI = "TELEFONE_HSM"
COL_ENRIQUECIDO  = "_ENRIQUECIDO"
# Aba onde ficam os dados tabulares (evita ler "sms"/"Emails" se estiverem 1.º no Excel)
ABA_PLANILHA_PRINCIPAL = "Principal"
ABA_DISPARO_HSM = "Disparo_HSM"

# Detecta colunas de email nas planilhas de enriquecimento
PADRAO_EMAIL_COL = re.compile(r"EMAIL", re.IGNORECASE)


def _normalizar_cpf(cpf) -> str:
    """
    11 digitos; zfill(11) a esquerda se faltarem.
    Trata int/float (Excel), strings com sufixo '.0' e leitura com dtype=str (etapa 2).
    """
    if cpf is None or (isinstance(cpf, float) and pd.isna(cpf)):
        return ""
    if isinstance(cpf, bool):
        return ""
    if isinstance(cpf, (int, float)):
        if isinstance(cpf, float) and not cpf == int(cpf):
            s = str(cpf)
        else:
            s = str(int(cpf))
    else:
        s = str(cpf).strip()
    if s.lower() in ("", "nan", "none", "nat"):
        return ""
    if s.endswith(".0"):
        b = s[:-2].lstrip("-")
        if b.isdigit():
            s = s[:-2]
    dig = re.sub(r"\D", "", s)
    if not dig:
        return ""
    if len(dig) > 11:
        dig = dig[-11:]
    if len(dig) < 11:
        return dig.zfill(11)
    return dig


def _normalizar_nome_cruzamento(nome) -> str:
    """
    Nome canonico para cruzamento P2-so-NOME ↔ Requerente: maiúsculas, sem acentos,
    espacos colapsados. Vazio se invalido.
    """
    if nome is None or (isinstance(nome, float) and pd.isna(nome)):
        return ""
    if isinstance(nome, bool):
        return ""
    s = str(nome).strip()
    if not s or s.lower() in ("nan", "none", "nat"):
        return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return " ".join(s.upper().split())


def _coluna_nome_na_principal(df: pd.DataFrame) -> str:
    """Coluna de nome do requerente (ou sinonimo) na planilha principal."""
    for cand in ("Requerente", "NOME", "Nome"):
        if cand in df.columns:
            return cand
    raise ValueError(
        "Para cruzamento por nome (planilha 2 sem CPF/CNPJ) a principal precisa de "
        f"alguma das colunas: Requerente, NOME ou Nome. Presentes: {list(df.columns)!r}"
    )


def _linha_ja_enriquecida_p2(val) -> bool:
    """Coluna _ENRIQUECIDO apos regravar a intermediaria (bool, 0/1, str, etc.)."""
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, float) and pd.isna(val):
        return False
    if isinstance(val, (int, float)):
        return int(val) != 0
    t = str(val).strip().lower()
    if t in ("true", "1", "yes", "verdadeiro", "t", "1.0"):
        return True
    if t in ("false", "0", "no", "falso", "f", "", "0.0"):
        return False
    return False


def _aba_parece_planilha_dados_principal(df: pd.DataFrame) -> bool:
    """Distingue a folha tabular das abas de explosao."""
    cols = list(df.columns)
    if COL_ENRIQUECIDO in cols:
        return True
    return any(str(c).startswith(f"{PREFIXO_TELEFONE}_") for c in cols)


def carregar_planilha_principal_de_workbook(caminho: str) -> pd.DataFrame:
    """
    Le a intermediaria/final garantindo os dados principais — nunca a primeira
    folha apenas por ordem (ex.: sms a frente por reordenacao manual no Excel).
    """
    xl = pd.ExcelFile(caminho)
    nomes_ok = xl.sheet_names
    for nome in (ABA_PLANILHA_PRINCIPAL, "Sheet1"):
        if nome in nomes_ok:
            df = pd.read_excel(caminho, sheet_name=nome, dtype=str)
            if _aba_parece_planilha_dados_principal(df) or len(df) > 0:
                return df
    for nome in nomes_ok:
        if nome.lower() in ("sms", "emails") or nome == ABA_DISPARO_HSM:
            continue
        df = pd.read_excel(caminho, sheet_name=nome, dtype=str)
        if _aba_parece_planilha_dados_principal(df):
            return df
    if nomes_ok:
        return pd.read_excel(caminho, sheet_name=nomes_ok[0], dtype=str)
    raise ValueError(f"Workbook sem folhas: {caminho!r}")


def _nome_coluna_cpf_vinculo_enriquecimento(df: pd.DataFrame) -> str | None:
    """Segunda coluna CPF (Excel duplicado). Pandas: `CPF.1` ou `cpf.1` conforme o rótulo."""
    if "CPF.1" in df.columns:
        return "CPF.1"
    if "cpf.1" in df.columns:
        return "cpf.1"
    return None


def _coluna_cpf_cruzamento_enriquecimento(
    df: pd.DataFrame, modelo: str | None = None
) -> str:
    """
    PRC CMP e PRC IMP: cruzamento P2/P3 com a segunda coluna CPF
    (`CPF.1` ou `cpf.1`). Sem modelo (etapa 2), usa a segunda se existir.
    """
    m = (modelo or "").strip().lower()
    v = _nome_coluna_cpf_vinculo_enriquecimento(df)
    if m in ("prc_cmp", "prc_imp") and v:
        return v
    if modelo is None and v:
        return v
    return "CPF"


def _coletar_contatos(
    df: pd.DataFrame, chave_normalizada: str, modo_merge_p2: str
) -> tuple[list[str], list[str]]:
    """
    Retorna (telefones, emails) nao vazios para uma chave na planilha de enriquecimento.

    modo_merge_p2: ``\"cpf\"`` usa ``_CPF_NORM``; ``\"nome\"`` usa ``_NOME_NORM``.
    """
    col_idx = "_CPF_NORM" if modo_merge_p2 == "cpf" else "_NOME_NORM"
    linhas = df[df[col_idx] == chave_normalizada]
    if linhas.empty:
        return [], []

    skip = {P2_COL_NOME, P2_COL_CPF, "_CPF_NORM", "_NOME_NORM"}
    colunas_dados = [c for c in df.columns if c not in skip]
    telefones, emails = [], []

    for _, row in linhas.iterrows():
        for col in colunas_dados:
            val = str(row[col]).strip() if pd.notna(row[col]) else ""
            if not val or val.lower() == "nan":
                continue
            if PADRAO_EMAIL_COL.search(col):
                emails.append(val)
            else:
                telefones.append(val)

    return telefones, emails


def _coletar_hsm_lemitti(
    df_p2: pd.DataFrame, chave_normalizada: str, modo_merge_p2: str
) -> list[tuple[str, bool]]:
    """Telefones HSM vindos só da concatenação das colunas Excel BA+BB (série `_HSM_BA_BB`)."""
    col_idx = "_CPF_NORM" if modo_merge_p2 == "cpf" else "_NOME_NORM"
    linhas = df_p2[df_p2[col_idx] == chave_normalizada]
    if linhas.empty:
        return []
    hsms: list[str] = []
    for _, row in linhas.iterrows():
        v = row.get("_HSM_BA_BB")
        s = str(v).strip() if pd.notna(v) else ""
        if s and s.lower() != "nan":
            hsms.append(s)
    return _deduplicar(hsms, is_red=True)


def _aplicar_destaque_hsm_na_planilha(
    ws, colunas_tel: list[str], colunas_hsm: list[str]
) -> None:
    """Realça em amarelo e fonte forte as colunas HSM e as células de TELEFONE_* iguais ao HSM."""
    cab = {cell.value: cell.column for cell in ws[1]}
    fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    fonte_h = Font(bold=True, color="0D47A1")
    col_hsm = [cab[c] for c in colunas_hsm if c in cab]
    col_tel = [cab[c] for c in colunas_tel if c in cab]

    def apenas_digitos(x) -> str:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return ""
        return re.sub(r"\D", "", str(x).strip())

    for r in range(2, ws.max_row + 1):
        alvo: set[str] = set()
        for ci in col_hsm:
            cell = ws.cell(row=r, column=ci)
            d = apenas_digitos(cell.value)
            if not d:
                continue
            alvo.add(d)
            if d.startswith("55") and len(d) > 2:
                alvo.add(d[2:])
            if not d.startswith("55") and len(d) >= 8:
                alvo.add("55" + d)

        for ci in col_hsm:
            cell = ws.cell(row=r, column=ci)
            if cell.value is not None and str(cell.value).strip():
                cell.fill = fill
                cell.font = fonte_h

        if not alvo:
            continue
        for ci in col_tel:
            cell = ws.cell(row=r, column=ci)
            d = apenas_digitos(cell.value)
            if not d:
                continue
            ok = d in alvo or (
                d.startswith("55") and len(d) > 2 and d[2:] in alvo
            ) or ("55" + d) in alvo
            if ok:
                cell.fill = fill
                cell.font = fonte_h


def _formatar_contato_55(telefone) -> str:
    """Prefixa 55 ao numero (apenas digitos)."""
    if telefone is None or (isinstance(telefone, float) and pd.isna(telefone)):
        return ""
    s = re.sub(r"\D", "", str(telefone).strip())
    if not s or s.lower() == "nan":
        return ""
    if s.startswith("55"):
        return s
    return "55" + s


def _formatar_nome_sms(requerente) -> str:
    """
    Primeiro e ultimo nome por extenso; nomes do meio apenas com inicial.
    Ex.: FILIPE NOBERTO DA SILVA JUSTINO -> Filipe N. D. S. Justino
    """
    if requerente is None or (isinstance(requerente, float) and pd.isna(requerente)):
        return ""
    nome = str(requerente).strip()
    if not nome or nome.lower() == "nan":
        return ""
    partes = nome.split()
    if not partes:
        return ""
    if len(partes) == 1:
        return partes[0].capitalize()
    if len(partes) == 2:
        return f"{partes[0].capitalize()} {partes[1].capitalize()}"
    primeira = partes[0].capitalize()
    ultima   = partes[-1].capitalize()
    meio     = [f"{p[0].upper()}." for p in partes[1:-1] if p]
    return " ".join([primeira] + meio + [ultima])


def _coluna_nome_disparo(df: pd.DataFrame) -> str | None:
    for cand in ("Requerente", "NOME", "Nome"):
        if cand in df.columns:
            return cand
    return None


def _coluna_processo_disparo(df: pd.DataFrame) -> str | None:
    if "Numero_de_Processo" in df.columns:
        return "Numero_de_Processo"
    if "Processo" in df.columns:
        return "Processo"
    return None


def _coluna_incidente_disparo(df: pd.DataFrame) -> str | None:
    for cand in ("Numero_do_Incidente", "Incidente"):
        if cand in df.columns:
            return cand
    return None


def _valor_cel_excel_py(v):
    """None para celula vazia / NaN; senao valor bruto."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return v


def _criar_aba_disparo_hsm(
    wb,
    df: pd.DataFrame,
    registros_hsm: list[list[tuple]],
) -> None:
    """
    Explosao por telefone HSM (BA+BB): Telefone HSM + Nome + Numero de Processo + Incidente.
    """
    n = len(df)
    if len(registros_hsm) != n:
        print(
            "     [AVISO] Aba Disparo_HSM omitida: inconsistencia entre DataFrame e registros_hsm."
        )
        return
    if sum(len(r) for r in registros_hsm) == 0:
        print("     Aba 'Disparo_HSM': omitida (nenhum telefone HSM).")
        return

    nome_c = _coluna_nome_disparo(df)
    proc_c = _coluna_processo_disparo(df)
    inc_c = _coluna_incidente_disparo(df)

    ws = wb.create_sheet(title=ABA_DISPARO_HSM)
    headers = ["Telefone HSM", "Nome", "Numero de Processo", "Incidente"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    fonte_h = Font(bold=True, color="0D47A1")

    row_excel = 2
    for df_idx, (_, row) in enumerate(df.iterrows()):
        for valor, _ in registros_hsm[df_idx]:
            sval = str(valor).strip() if valor is not None else ""
            if not sval or sval.lower() == "nan":
                continue
            c_tel = ws.cell(row=row_excel, column=1, value=sval)
            c_tel.fill = fill
            c_tel.font = fonte_h

            nome_v = _valor_cel_excel_py(row[nome_c]) if nome_c else None
            proc_v = _valor_cel_excel_py(row[proc_c]) if proc_c else None
            inc_v = _valor_cel_excel_py(row[inc_c]) if inc_c else None
            ws.cell(row=row_excel, column=2, value=nome_v)
            ws.cell(row=row_excel, column=3, value=proc_v)
            ws.cell(row=row_excel, column=4, value=inc_v)
            row_excel += 1

    print(f"     Aba 'Disparo_HSM': {row_excel - 2} linha(s).")


def _emitir_relatorio_blacklist(detalhes: list[dict], pasta_resultados: str) -> None:
    """Grava CSV e lista um resumo no terminal."""
    os.makedirs(pasta_resultados, exist_ok=True)
    caminho = os.path.join(pasta_resultados, "blacklist_bloqueios.csv")
    cols = ["tipo_bloqueio", "cpf", "requerente", "numero_processo", "valor_removido"]
    if detalhes:
        df_rel = pd.DataFrame(detalhes).reindex(columns=cols)
    else:
        df_rel = pd.DataFrame(columns=cols)
    df_rel.to_csv(caminho, index=False, encoding="utf-8-sig")
    print(f"     Arquivo com todos os bloqueios: {caminho}")
    if not detalhes:
        print("     (Nenhum item bloqueado nesta execucao.)")
        return
    lim = 50
    print(f"     Resumo no terminal (ate {lim} linhas):")
    for d in detalhes[:lim]:
        req = str(d.get("requerente") or "")[:45]
        proc = str(d.get("numero_processo") or "")
        cpf = str(d.get("cpf") or "")
        print(
            f"       [{d.get('tipo_bloqueio')}] processo={proc} | cpf={cpf} | {req} "
            f"-> {str(d.get('valor_removido'))[:70]}"
        )
    if len(detalhes) > lim:
        print(f"       ... e mais {len(detalhes) - lim} linha(s) no CSV acima.")


def _deduplicar(valores: list[str], is_red: bool) -> list[tuple[str, bool]]:
    """Remove duplicatas mantendo a ordem, retorna lista de (valor, is_red)."""
    vistos, resultado = set(), []
    for v in valores:
        if v not in vistos:
            vistos.add(v)
            resultado.append((v, is_red))
    return resultado


def _preencher_colunas(
    df: pd.DataFrame, registros: list[list[tuple]], prefixo: str
) -> tuple[pd.DataFrame, list[str]]:
    """Adiciona colunas PREFIXO_1, PREFIXO_2, ... ao DataFrame com os valores."""
    df.reset_index(drop=True, inplace=True)
    max_itens = max((len(r) for r in registros), default=0)
    colunas   = [f"{prefixo}_{i+1}" for i in range(max_itens)]

    for col in colunas:
        if col not in df.columns:
            df[col] = pd.NA

    n_df = len(df)
    if len(registros) != n_df:
        raise ValueError(
            f"_preencher_colunas ({prefixo}): DataFrame tem {n_df} linhas, "
            f"registros tem {len(registros)}."
        )

    for pos, itens in enumerate(registros):
        for j, (val, _) in enumerate(itens):
            col = colunas[j]
            df.iloc[pos, df.columns.get_loc(col)] = val

    return df, colunas


def _aplicar_cores(
    ws,
    registros_tel: list, colunas_tel: list,
    registros_email: list, colunas_email: list,
) -> None:
    """Aplica cor vermelha (planilha 2) ou preta (planilha 3) nas celulas."""
    cabecalho      = {cell.value: cell.column for cell in ws[1]}
    fonte_vermelha = Font(color="FF0000")
    fonte_normal   = Font(color="000000")

    for grupos in [(registros_tel, colunas_tel), (registros_email, colunas_email)]:
        registros, colunas = grupos
        for row_excel, itens in enumerate(registros, start=2):
            for j, (_, is_red) in enumerate(itens):
                col_idx = cabecalho.get(colunas[j])
                if col_idx:
                    ws.cell(row=row_excel, column=col_idx).font = (
                        fonte_vermelha if is_red else fonte_normal
                    )


def _criar_aba_explosao(
    wb,
    df: pd.DataFrame,
    colunas_contato: list[str],
    registros: list[list[tuple]],
    nome_aba: str,
    nome_coluna: str,
    sms_extras: bool = False,
) -> None:
    """
    Cria uma aba no workbook onde cada linha e multiplicada pela quantidade
    de valores presentes nas colunas_contato. Uma linha por valor.
    Aplica cor vermelha na celula de contato quando o valor vem da planilha 2.

    Args:
        wb:               Workbook openpyxl ja aberto.
        df:               DataFrame principal.
        colunas_contato:  Colunas a explodir (ex: TELEFONE_1, TELEFONE_2...).
        registros:        Lista de [(valor, is_red), ...] por linha — fonte da cor.
        nome_aba:         Nome da aba a criar (ex: "sms", "Emails").
        nome_coluna:      Nome da coluna unica de contato na aba (ex: "TELEFONE", "EMAIL").
        sms_extras:       Se True (aba sms), adiciona ao final: Contato (55...), Nome (formatado), Processo.
    """
    colunas_base = [c for c in df.columns if c not in colunas_contato]
    col_contato_idx = len(colunas_base) + 1  # indice (1-based) da coluna de contato na aba

    fonte_vermelha = Font(color="FF0000")
    fonte_normal   = Font(color="000000")

    ws = wb.create_sheet(title=nome_aba)

    extras = ["Contato", "Nome", "Processo"] if sms_extras else []
    cabecalho = colunas_base + [nome_coluna] + extras

    # Cabecalho
    for col_idx, col_name in enumerate(cabecalho, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    row_excel = 2
    for df_idx, (_, row) in enumerate(df.iterrows()):
        itens = registros[df_idx]  # [(valor, is_red), ...]
        for valor, is_red in itens:
            # Dados base
            for col_idx, col_name in enumerate(colunas_base, start=1):
                val = row[col_name]
                ws.cell(row=row_excel, column=col_idx, value=val if pd.notna(val) else None)
            # Coluna de contato com cor
            cell = ws.cell(row=row_excel, column=col_contato_idx, value=valor)
            cell.font = fonte_vermelha if is_red else fonte_normal
            if sms_extras:
                c = col_contato_idx + 1
                proc = row.get("Numero_de_Processo")
                ws.cell(row=row_excel, column=c,     value=_formatar_contato_55(valor))
                ws.cell(row=row_excel, column=c + 1, value=_formatar_nome_sms(row.get("Requerente")))
                ws.cell(row=row_excel, column=c + 2, value=None if proc is None or pd.isna(proc) else proc)
            row_excel += 1

    total_linhas = row_excel - 2
    print(f"     Aba '{nome_aba}': {total_linhas} linhas geradas a partir de {len(df)} registros")


def _salvar_com_cores(
    df: pd.DataFrame,
    registros_tel: list, colunas_tel: list,
    registros_email: list, colunas_email: list,
    caminho: str,
    colunas_hsm: list | None = None,
    registros_hsm: list | None = None,
) -> None:
    """Salva o DataFrame em Excel com cores e cria as abas 'sms', 'Emails' e 'Disparo_HSM'."""
    df.to_excel(caminho, index=False, sheet_name=ABA_PLANILHA_PRINCIPAL)
    wb = load_workbook(caminho)
    ws = wb[ABA_PLANILHA_PRINCIPAL]
    _aplicar_cores(ws, registros_tel, colunas_tel, registros_email, colunas_email)

    chsm = colunas_hsm or []
    if chsm:
        _aplicar_destaque_hsm_na_planilha(ws, colunas_tel, chsm)

    # Colunas base = todas exceto telefone, email e HSM Lemitti
    colunas_base = [
        c for c in df.columns
        if not c.startswith(f"{PREFIXO_TELEFONE}_")
        and not c.startswith(f"{PREFIXO_EMAIL}_")
        and not c.startswith(f"{PREFIXO_HSM_LEMITTI}_")
    ]

    print(f"\n     Gerando abas de explosao...")
    _criar_aba_explosao(
        wb, df[colunas_base + colunas_tel], colunas_tel, registros_tel,
        nome_aba="sms", nome_coluna="TELEFONE", sms_extras=True,
    )
    _criar_aba_explosao(wb, df[colunas_base + colunas_email], colunas_email, registros_email, nome_aba="Emails", nome_coluna="EMAIL")

    if registros_hsm is not None:
        _criar_aba_disparo_hsm(wb, df, registros_hsm)

    wb.save(caminho)


def exportar_bytes_prc_lemitti_hsm(
    caminho_principal: str,
    caminho_p2: str,
    modelo: str = "prc_tjsp",
) -> bytes:
    """
    Uma folha Excel: planilha principal + telefones Lemitti (DDD+FONE) + colunas
    ``TELEFONE_HSM_*`` (concatenação colunas Excel BA+BB do CSV Lemitti).
    Células HSM e qualquer ``TELEFONE_*`` com o mesmo número são realçadas.
    """
    criar_banco_e_tabelas()

    df_main = processar_planilha_principal(caminho_principal, modelo=modelo)
    df_p2, modo_merge_p2 = processar_enriquecimento_contatos(caminho_p2)

    serie_hsm = ler_serie_telefone_concat_colunas_excel(caminho_p2)
    df_p2 = df_p2.reset_index(drop=True)
    serie_hsm = serie_hsm.reset_index(drop=True)
    if len(serie_hsm) != len(df_p2):
        if len(serie_hsm) > len(df_p2):
            serie_hsm = serie_hsm.iloc[: len(df_p2)].reset_index(drop=True)
        else:
            serie_hsm = pd.concat(
                [
                    serie_hsm,
                    pd.Series([""] * (len(df_p2) - len(serie_hsm)), dtype=object),
                ],
                ignore_index=True,
            )
    df_p2["_HSM_BA_BB"] = serie_hsm.astype(str).fillna("").values

    col_cpf_x = _coluna_cpf_cruzamento_enriquecimento(df_main, modelo)
    if modo_merge_p2 == "cpf":
        if col_cpf_x not in df_main.columns:
            raise KeyError(
                f"Coluna de CPF inexistente: {col_cpf_x!r} (modelo={modelo!r}). Colunas: {list(df_main.columns)}"
            )
        df_main["_CPF_NORM"] = df_main[col_cpf_x].apply(_normalizar_cpf)
        df_p2["_CPF_NORM"] = df_p2[P2_COL_CPF].apply(_normalizar_cpf)
    else:
        col_nom = _coluna_nome_na_principal(df_main)
        if col_cpf_x not in df_main.columns:
            raise KeyError(
                f"Coluna de CPF inexistente: {col_cpf_x!r} (modelo={modelo!r}). "
                "Continua necessaria na principal mesmo no cruzamento por nome. "
                f"Colunas: {list(df_main.columns)}"
            )
        df_main["_CPF_NORM"] = df_main[col_cpf_x].apply(_normalizar_cpf)
        df_main["_NOME_MERGE"] = df_main[col_nom].apply(_normalizar_nome_cruzamento)
        df_p2["_NOME_NORM"] = df_p2[P2_COL_NOME].apply(_normalizar_nome_cruzamento)

    registros_tel: list = []
    registros_email: list = []
    registros_hsm: list = []
    df_p2_somente_contatos = df_p2.drop(columns=["_HSM_BA_BB"], errors="ignore")

    for _, row in df_main.iterrows():
        chave = row["_CPF_NORM"] if modo_merge_p2 == "cpf" else row["_NOME_MERGE"]
        fones, emails = _coletar_contatos(df_p2_somente_contatos, chave, modo_merge_p2)
        hsm_linha = _coletar_hsm_lemitti(df_p2, chave, modo_merge_p2)

        if fones or emails:
            registros_tel.append(_deduplicar(fones, is_red=True))
            registros_email.append(_deduplicar(emails, is_red=True))
        else:
            registros_tel.append([])
            registros_email.append([])
        registros_hsm.append(hsm_linha if hsm_linha else [])

    bl = carregar_blacklist()
    registros_tel, registros_email, *_rest = filtrar_registros_por_blacklist(
        df_main, registros_tel, registros_email, bl
    )
    registros_hsm = filtrar_hsm_por_blacklist(df_main, registros_hsm, bl)

    _drop_aux = ["_CPF_NORM"]
    if modo_merge_p2 == "nome":
        _drop_aux.append("_NOME_MERGE")
    df_main.drop(columns=_drop_aux, inplace=True)

    df_main, colunas_tel = _preencher_colunas(df_main, registros_tel, PREFIXO_TELEFONE)
    df_main, colunas_email = _preencher_colunas(df_main, registros_email, PREFIXO_EMAIL)
    df_main, colunas_hsm = _preencher_colunas(df_main, registros_hsm, PREFIXO_HSM_LEMITTI)

    buf = io.BytesIO()
    df_main.to_excel(buf, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    _aplicar_cores(ws, registros_tel, colunas_tel, registros_email, colunas_email)
    _aplicar_destaque_hsm_na_planilha(ws, colunas_tel, colunas_hsm)

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# ETAPA 1 — Enriquecimento com planilha 2
# Gera apenas:
#   - INTERMEDIARIA.xlsx            (resultado parcial com telefones e emails separados)
#   - cpfs_nao_encontrados_p2.csv   (CPFs para emitir a planilha 3)
# ══════════════════════════════════════════════════════════════════════════════
def etapa1_enriquecer_com_p2(
    caminho_principal: str,
    caminho_p2: str,
    caminho_saida_intermediaria: str,
    caminho_csv_nao_encontrados: str,
    caminho_blacklist_txt: str = "blacklist.txt",
    modelo: str = "prc_tjsp",
) -> None:
    """
    ETAPA 1 — Processa a planilha principal com os dados da planilha 2 (contatos).
    Telefones e emails sao separados em grupos de colunas distintos.
    Tudo processado em memoria. Salva apenas a intermediaria e o CSV de nao encontrados.
    """
    print("\n[0/3] Carregando blacklist do banco...")
    criar_banco_e_tabelas()

    print("\n[1/3] Processando planilha principal...")
    df_main = processar_planilha_principal(caminho_principal, modelo=modelo)

    print("\n[2/3] Processando planilha 2 (contatos)...")
    df_p2, modo_merge_p2 = processar_enriquecimento_contatos(caminho_p2)

    col_cpf_x = _coluna_cpf_cruzamento_enriquecimento(df_main, modelo)
    if modo_merge_p2 == "cpf":
        print(
            f"\n[3/3] Cruzando com planilha 2 por documento (CPF principal: {col_cpf_x!r})..."
        )
        if col_cpf_x not in df_main.columns:
            raise KeyError(
                f"Coluna de CPF inexistente: {col_cpf_x!r} (modelo={modelo!r}). Colunas: {list(df_main.columns)}"
            )
        df_main["_CPF_NORM"] = df_main[col_cpf_x].apply(_normalizar_cpf)
        df_p2["_CPF_NORM"] = df_p2[P2_COL_CPF].apply(_normalizar_cpf)
    else:
        col_nom = _coluna_nome_na_principal(df_main)
        print(
            f"\n[3/3] Cruzando com planilha 2 por NOME ({col_nom!r} × "
            f"{P2_COL_NOME!r})..."
        )
        if col_cpf_x not in df_main.columns:
            raise KeyError(
                f"Coluna de CPF inexistente: {col_cpf_x!r} (modelo={modelo!r}). "
                "Continua necessaria na principal mesmo no cruzamento por nome "
                "(CSV de nao encontrados e etapas seguintes)."
                f" Colunas: {list(df_main.columns)}"
            )
        df_main["_CPF_NORM"] = df_main[col_cpf_x].apply(_normalizar_cpf)
        df_main["_NOME_MERGE"] = df_main[col_nom].apply(_normalizar_nome_cruzamento)
        df_p2["_NOME_NORM"] = df_p2[P2_COL_NOME].apply(_normalizar_nome_cruzamento)

    serie_hsm = ler_serie_telefone_concat_colunas_excel(caminho_p2)
    df_p2 = df_p2.reset_index(drop=True)
    serie_hsm = serie_hsm.reset_index(drop=True)
    if len(serie_hsm) != len(df_p2):
        if len(serie_hsm) > len(df_p2):
            serie_hsm = serie_hsm.iloc[: len(df_p2)].reset_index(drop=True)
        else:
            serie_hsm = pd.concat(
                [
                    serie_hsm,
                    pd.Series([""] * (len(df_p2) - len(serie_hsm)), dtype=object),
                ],
                ignore_index=True,
            )
    df_p2["_HSM_BA_BB"] = serie_hsm.astype(str).fillna("").values
    df_p2_somente_contatos = df_p2.drop(columns=["_HSM_BA_BB"], errors="ignore")

    registros_tel        = []
    registros_email      = []
    registros_hsm: list[list[tuple[str, bool]]] = []
    cpfs_nao_encontrados = []

    for _, row in df_main.iterrows():
        if modo_merge_p2 == "cpf":
            chave = row["_CPF_NORM"]
        else:
            chave = row["_NOME_MERGE"]
        fones, emails = _coletar_contatos(df_p2_somente_contatos, chave, modo_merge_p2)
        hsm_linha = _coletar_hsm_lemitti(df_p2, chave, modo_merge_p2)
        registros_hsm.append(hsm_linha if hsm_linha else [])

        if fones or emails:
            registros_tel.append(_deduplicar(fones, is_red=True))
            registros_email.append(_deduplicar(emails, is_red=True))
        else:
            registros_tel.append([])
            registros_email.append([])
            cpfs_nao_encontrados.append({"CPF": row[col_cpf_x]})

    pd.DataFrame(cpfs_nao_encontrados).to_csv(caminho_csv_nao_encontrados, index=False)

    df_main[COL_ENRIQUECIDO] = [bool(t or e) for t, e in zip(registros_tel, registros_email)]
    _drop_aux = ["_CPF_NORM"]
    if modo_merge_p2 == "nome":
        _drop_aux.append("_NOME_MERGE")
    df_main.drop(columns=_drop_aux, inplace=True)

    # ── Blacklist antes de gravar TELEFONE / EMAIL / TELEFONE_HSM (valores === listas filtradas)
    print("\n     Aplicando blacklist...")
    bl = carregar_blacklist()
    registros_tel, registros_email, p_bloq, t_bloq, e_bloq, bl_detalhes = filtrar_registros_por_blacklist(
        df_main, registros_tel, registros_email, bl
    )
    registros_hsm = filtrar_hsm_por_blacklist(df_main, registros_hsm, bl)
    print(
        f"     Blacklist: {p_bloq} pessoa(s) sem contato | "
        f"{t_bloq} telefone(s) | {e_bloq} email(s) removido(s)."
    )
    _emitir_relatorio_blacklist(bl_detalhes, os.path.dirname(caminho_saida_intermediaria))

    df_main, colunas_tel = _preencher_colunas(df_main, registros_tel, PREFIXO_TELEFONE)
    df_main, colunas_email = _preencher_colunas(df_main, registros_email, PREFIXO_EMAIL)
    df_main, colunas_hsm = _preencher_colunas(df_main, registros_hsm, PREFIXO_HSM_LEMITTI)

    _salvar_com_cores(
        df_main,
        registros_tel,
        colunas_tel,
        registros_email,
        colunas_email,
        caminho_saida_intermediaria,
        colunas_hsm=colunas_hsm,
        registros_hsm=registros_hsm,
    )

    # ── Banco de dados — apenas log da execucao, sem incrementar contadores ──
    # Counts e ultimo_processamento so sao atualizados no disparo final (etapa2)
    print("\n     Registrando execucao no banco...")
    encontrados = len(df_main) - len(cpfs_nao_encontrados)
    registrar_execucao(
        etapa=1,
        arquivo_principal=caminho_principal,
        arquivo_p2=caminho_p2,
        total_registros=len(df_main),
        total_enriquecidos_p2=encontrados,
        total_sem_contato=len(cpfs_nao_encontrados),
    )

    print(f"\n[OK] Etapa 1 concluida.")
    print(f"     CPFs encontrados na P2    : {encontrados}")
    print(f"     CPFs NAO encontrados na P2: {len(cpfs_nao_encontrados)}")
    print(f"     Planilha intermediaria    : {caminho_saida_intermediaria}")
    print(f"     CSV nao encontrados       : {caminho_csv_nao_encontrados}")
    print(f"\n  >> Emita a planilha 3 com os CPFs do CSV e execute a Etapa 2.")


# ══════════════════════════════════════════════════════════════════════════════
# ETAPA 2 — Enriquecimento com planilha 3
# Retoma a partir da intermediaria. Gera apenas:
#   - FINAL.xlsx
# ══════════════════════════════════════════════════════════════════════════════
def etapa2_enriquecer_com_p3(
    caminho_intermediaria: str,
    caminho_p3: str,
    caminho_saida_final: str,
    caminho_blacklist_txt: str = None,
) -> None:
    """
    ETAPA 2 — Completa o enriquecimento com a planilha 3 (relacionados).
    Deve ser executada APOS a Etapa 1 e APOS emitir a planilha 3.
    Telefones e emails sao separados em grupos de colunas distintos.
    """
    criar_banco_e_tabelas()

    print("\n[1/3] Carregando planilha intermediaria...")
    df_main = carregar_planilha_principal_de_workbook(caminho_intermediaria)
    df_main.columns = pd.Index([str(c).strip() for c in df_main.columns])
    print(f"     Linhas carregadas: {len(df_main)}")

    print("\n[2/3] Processando planilha 3 (relacionados)...")
    df_p3 = processar_enriquecimento_relacionados(caminho_p3)

    col_cpf_x = _coluna_cpf_cruzamento_enriquecimento(df_main, modelo=None)
    print(
        f"\n[3/3] Cruzando CPFs nao enriquecidos com planilha 3 (chave: {col_cpf_x})..."
    )
    if col_cpf_x not in df_main.columns:
        raise KeyError(
            f"Coluna de CPF inexistente apos reabrir a intermediaria: {col_cpf_x!r}. Colunas: {list(df_main.columns)}"
        )
    df_main["_CPF_NORM"] = df_main[col_cpf_x].apply(_normalizar_cpf)
    df_p3["_CPF_NORM"]   = df_p3["CPF"].apply(_normalizar_cpf)

    colunas_tel_exist   = [c for c in df_main.columns if str(c).startswith(f"{PREFIXO_TELEFONE}_")]
    colunas_email_exist = [c for c in df_main.columns if str(c).startswith(f"{PREFIXO_EMAIL}_")]
    colunas_hsm_exist   = [
        c for c in df_main.columns if str(c).startswith(f"{PREFIXO_HSM_LEMITTI}_")
    ]

    # Reconstroi registros preservando dados da etapa 1 (vermelho) + HSM (Lemitti)
    registros_tel: list[list[tuple[str, bool]]]   = []
    registros_email: list[list[tuple[str, bool]]] = []
    registros_hsm: list[list[tuple[str, bool]]]   = []

    for _, row in df_main.iterrows():
        tel_row = []
        for col in colunas_tel_exist:
            val = str(row[col]).strip() if pd.notna(row[col]) and str(row[col]).strip() != "nan" else ""
            if val:
                tel_row.append((val, True))
        registros_tel.append(tel_row)

        email_row = []
        for col in colunas_email_exist:
            val = str(row[col]).strip() if pd.notna(row[col]) and str(row[col]).strip() != "nan" else ""
            if val:
                email_row.append((val, True))
        registros_email.append(email_row)

        hsm_row = []
        for col in colunas_hsm_exist:
            val = str(row[col]).strip() if pd.notna(row[col]) and str(row[col]).strip() != "nan" else ""
            if val:
                hsm_row.append((val, True))
        registros_hsm.append(hsm_row)

    enriquecidos_p3 = 0
    for pos in range(len(df_main)):
        row = df_main.iloc[pos]
        if _linha_ja_enriquecida_p2(row.get(COL_ENRIQUECIDO, False)):
            continue

        fones, emails = _coletar_contatos(df_p3, row["_CPF_NORM"], "cpf")
        if fones or emails:
            registros_tel[pos]   = _deduplicar(fones,  is_red=False)
            registros_email[pos] = _deduplicar(emails, is_red=False)
            enriquecidos_p3 += 1

    print(f"     CPFs enriquecidos via P3: {enriquecidos_p3}")

    df_main.drop(
        columns=["_CPF_NORM", COL_ENRIQUECIDO]
        + colunas_tel_exist
        + colunas_email_exist
        + colunas_hsm_exist,
        inplace=True,
        errors="ignore",
    )

    # Metricas antes do cooldown (para o resumo final nao confundir 0 linhas com "sem dados")
    n_pre_cd = len(df_main)
    explosao_sms_pre_cd = sum(len(t) for t in registros_tel)
    explosao_email_pre_cd = sum(len(e) for e in registros_email)
    com_contato_pre_cd = sum(
        1 for t, e in zip(registros_tel, registros_email) if t or e
    )
    enriquecidos_p2_pre_cd = sum(
        1 for t, e in zip(registros_tel, registros_email)
        if any(is_red for _, is_red in t + e)
    )
    enriquecidos_p3_pre_cd = sum(
        1 for t, e in zip(registros_tel, registros_email)
        if (t or e) and not any(is_red for _, is_red in t + e)
    )

    # ── Cooldown — remove CPFs processados nos ultimos 14 dias ───────────────
    if _pular_cooldown_etapa2():
        total_cooldown = 0
        print("\n     Cooldown: desativado (env EDA_SKIP_COOLDOWN=1, etc.). Nenhuma linha removida.")
    else:
        dias_cd = _dias_cooldown_etapa2()
        print(f"\n     Verificando cooldown ({dias_cd} dias)...")
        cpfs_cooldown = buscar_cpfs_cooldown(dias=dias_cd)
        if cpfs_cooldown:
            col_cd = _coluna_cpf_cruzamento_enriquecimento(df_main, modelo=None)
            if col_cd not in df_main.columns:
                raise KeyError(
                    f"Coluna de CPF inexistente no cooldown: {col_cd!r}. Colunas: {list(df_main.columns)}"
                )
            df_main["_CPF_NORM"] = df_main[col_cd].apply(_normalizar_cpf)
            mascara_cd_arr = df_main["_CPF_NORM"].isin(cpfs_cooldown).to_numpy()
            n_match = int(mascara_cd_arr.sum())
            n_total = len(df_main)
            total_cooldown = 0
            cooldown_bloqueou_todos = (
                n_match >= n_total and n_total > 0 and bool(mascara_cd_arr.all())
            )
            if cooldown_bloqueou_todos and not _cooldown_permitir_planilha_totalmente_filtrada():
                total_cooldown = 0
                print(
                    "     Cooldown: todos os registros coincidiriam com a janela; "
                    "o lote seria inteiromente removido e o FINAL ficaria vazio. "
                    "Mantendo todas as linhas nesta execucao. "
                    "Para desativar o filtro: EDA_SKIP_COOLDOWN=1 ou EDA_COOLDOWN_DIAS=0; "
                    "para permitir FINAL vazio: EDA_COOLDOWN_PERMITIR_PLANILHA_VAZIA=1."
                )
            elif n_match:
                pos_manter      = [
                    i for i in range(n_total) if not bool(mascara_cd_arr[i])
                ]
                removed = n_match
                registros_tel   = [registros_tel[i] for i in pos_manter]
                registros_email = [registros_email[i] for i in pos_manter]
                registros_hsm   = [registros_hsm[i] for i in pos_manter]
                df_main         = df_main.iloc[pos_manter].reset_index(drop=True)
                total_cooldown = removed
                print(
                    f"     Cooldown: {removed} registro(s) removido(s) da planilha final."
                )
            else:
                print("     Cooldown: nenhum registro em cooldown.")
            df_main.drop(columns=["_CPF_NORM"], inplace=True, errors="ignore")
        else:
            total_cooldown = 0
            print("     Cooldown: nenhum historico encontrado.")

    # ── Blacklist — filtra antes de gerar as abas ─────────────────────────────
    print("\n     Aplicando blacklist...")
    bl = carregar_blacklist()
    registros_tel, registros_email, p_bloq, t_bloq, e_bloq, bl_detalhes = filtrar_registros_por_blacklist(
        df_main, registros_tel, registros_email, bl
    )
    registros_hsm = filtrar_hsm_por_blacklist(df_main, registros_hsm, bl)
    print(
        f"     Blacklist: {p_bloq} pessoa(s) sem contato | "
        f"{t_bloq} telefone(s) | {e_bloq} email(s) removido(s)."
    )
    _emitir_relatorio_blacklist(bl_detalhes, os.path.dirname(caminho_saida_final))

    df_main, colunas_tel   = _preencher_colunas(df_main, registros_tel,   PREFIXO_TELEFONE)
    df_main, colunas_email = _preencher_colunas(df_main, registros_email, PREFIXO_EMAIL)
    df_main, colunas_hsm   = _preencher_colunas(df_main, registros_hsm,   PREFIXO_HSM_LEMITTI)

    _salvar_com_cores(
        df_main,
        registros_tel,
        colunas_tel,
        registros_email,
        colunas_email,
        caminho_saida_final,
        colunas_hsm=colunas_hsm,
        registros_hsm=registros_hsm,
    )

    # ── Banco de dados ────────────────────────────────────────────────────────
    print("\n     Salvando no banco de dados...")
    total_enriquecidos = sum(1 for t, e in zip(registros_tel, registros_email) if t or e)
    id_execucao = registrar_execucao(
        etapa=2,
        arquivo_p3=caminho_p3,
        total_registros=len(df_main),
        total_enriquecidos_p3=enriquecidos_p3,
        total_sem_contato=len(df_main) - total_enriquecidos,
    )
    mapa_ids = salvar_processos(df_main, id_execucao)
    salvar_contatos(df_main, registros_tel, registros_email, mapa_ids, id_execucao)
    salvar_disparo_hsm(df_main, registros_hsm, mapa_ids, id_execucao)

    total_sem_contato  = len(df_main) - total_enriquecidos
    total_sms_gerados  = sum(len(t) for t in registros_tel)
    total_email_gerado = sum(len(e) for e in registros_email)

    # Enriquecidos so pela P2, so pela P3, e por ambas (nao se aplica aqui pois e sequencial)
    enriquecidos_p2 = sum(
        1 for t, e in zip(registros_tel, registros_email)
        if any(is_red for _, is_red in t + e)
    )
    enriquecidos_p3_final = sum(
        1 for t, e in zip(registros_tel, registros_email)
        if (t or e) and not any(is_red for _, is_red in t + e)
    )

    taxa_enriquecimento = (total_enriquecidos / len(df_main) * 100) if len(df_main) else 0.0

    print(f"\n[OK] Planilha final salva em: {caminho_saida_final}")

    sep  = "=" * 55
    div  = "-" * 53
    data = __import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')

    print(f"\n{sep}")
    print(f"  RESUMO DO PROCESSAMENTO DIARIO - {data}")
    print(sep)
    print(f"  Linhas principais (pre-cooldown)     : {n_pre_cd}")
    print(f"  Removidas por cooldown (< 14 dias) : {total_cooldown}")
    print(f"  Linhas na planilha FINAL           : {len(df_main)}")
    if len(df_main) == 0 and total_cooldown > 0:
        print(
            "  (*) Nenhuma linha exportada: todo o lote coincidiu com CPF ja "
            "processado no periodo de cooldown da base."
        )
    elif len(df_main) == 0 and total_cooldown == 0:
        print(
            "  (*) Planilha vazia: verifique entrada/intermediaria (sem dados validos)."
        )
    print(f"  Com ao menos um contato (FINAL)   : {total_enriquecidos}")
    if total_cooldown or n_pre_cd != len(df_main):
        print(
            f"  Com ao menos um contato (pre-CD)  : {com_contato_pre_cd} "
            f"(referencia antes do cooldown)"
        )
    print(f"  {div}")
    print(f"  Enriquecidos Lemitti P2 (FINAL)    : {enriquecidos_p2}")
    print(f"  Enriquecidos Assertiva P3 (FINAL)  : {enriquecidos_p3_final}")
    if total_cooldown or n_pre_cd != len(df_main):
        print(f"  Enriquecidos P2 (pre-cooldown)     : {enriquecidos_p2_pre_cd}")
        print(f"  Enriquecidos P3 (pre-cooldown)     : {enriquecidos_p3_pre_cd}")
    print(f"  Sem contato na FINAL               : {total_sem_contato}")
    print(f"  Taxa enriquecimento (sobre FINAL)  : {taxa_enriquecimento:.1f}%")
    print(f"  {div}")
    print(f"  Linhas aba SMS (explosao, FINAL)   : {total_sms_gerados}")
    print(f"  Linhas aba Emails (explosao, FINAL): {total_email_gerado}")
    if total_cooldown or explosao_sms_pre_cd != total_sms_gerados:
        print(
            f"  Referencia explosao SMS (pre-CD)   : {explosao_sms_pre_cd} linhas"
        )
    if total_cooldown or explosao_email_pre_cd != total_email_gerado:
        print(
            f"  Referencia explosao Email (pre-CD) : {explosao_email_pre_cd} linhas"
        )
    print(f"  {div}")
    print("  Blacklist (MySQL EDA_MYSQL_DATABASE)")
    print(f"    Pessoas bloqueadas           : {p_bloq}")
    print(f"    Telefones bloqueados         : {t_bloq}")
    print(f"    Emails bloqueados            : {e_bloq}")
    print(f"  {div}")
    print(f"  Colunas reservadas telefone    : {len(colunas_tel)}")
    print(f"  Colunas reservadas email       : {len(colunas_email)}")
    print("  (com planilha vazia, o Excel mantem a estrutura de colunas.)")
    print(f"  {div}")
    print(f"  Vermelho = Lemitti (P2)     | Preto = Assertiva (P3)")
    print(f"{sep}\n")


# ══════════════════════════════════════════════════════════════════════════════
# Execucao direta (sem interface web) — ajuste os caminhos se necessario
# Normalmente rodado via: python app.py  (interface web)
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import os
    BASE = os.path.dirname(os.path.dirname(__file__))  # raiz do projeto

    # ── Rode ETAPA 1 primeiro ─────────────────────────────────────────────
    etapa1_enriquecer_com_p2(
        caminho_principal           = os.path.join(BASE, "Entrada",    "principal.xlsx"),
        caminho_p2                  = os.path.join(BASE, "Entrada",    "enriquecimento_lemitti.csv"),
        caminho_saida_intermediaria = os.path.join(BASE, "Resultados", "INTERMEDIARIA.xlsx"),
        caminho_csv_nao_encontrados = os.path.join(BASE, "Resultados", "cpfs_nao_encontrados_p2.csv"),
        caminho_blacklist_txt       = os.path.join(BASE, "blacklist.txt"),
    )

    # ── Depois de emitir a planilha 3, rode ETAPA 2 ───────────────────────
    # etapa2_enriquecer_com_p3(
    #     caminho_intermediaria = os.path.join(BASE, "Resultados", "INTERMEDIARIA.xlsx"),
    #     caminho_p3            = os.path.join(BASE, "Entrada",    "enriquecimento_assertiva.csv"),
    #     caminho_saida_final   = os.path.join(BASE, "Resultados", "DD-MM-AAAA PRC TJSP FINAL.xlsx"),
    #     caminho_blacklist_txt = os.path.join(BASE, "blacklist.txt"),
    # )
