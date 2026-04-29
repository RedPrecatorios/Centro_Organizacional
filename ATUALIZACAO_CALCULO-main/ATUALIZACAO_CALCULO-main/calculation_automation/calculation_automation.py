# -*- coding: utf-8 -*-
"""
Automação de planilha em Linux: openpyxl (sem Microsoft Excel / xlwings).

- ``load_workbook(..., keep_vba=True)`` preserva o .xlsm; não se executa VBA.
- openpyxl **não recalcula** fórmulas; o valor lido (ex. O313) vem do **cache** do
  ficheiro, salvo se o ficheiro for recalculado noutro ambiente.
- N306:O313: sincronizados com ``memoria_calculo`` (``db_handler.memoria_range_sync``)
  se ``MEMORIA_MYSQL_*`` estiver no .env; com ``MEMORIA_LIBREOFFICE=1`` (padrão) tenta
  recálculo com LibreOffice headless antes de ler as células.
"""
import os
import re
import time
import locale
import shutil
import zipfile
import traceback
import unicodedata
from datetime import datetime

import openpyxl
import pytz
from openpyxl import load_workbook
from colorama import Fore, Style

from db_handler.db_handler import DBHandler
from db_handler.memoria_range_sync import (
    libreoffice_recalc_export_xlsx,
    load_merged_memoria_valores,
    memoria_recalc_wanted,
    sync_memoria_calculo_to_db,
    total_liquido_arredondado,
)
from txt_handler.txt_handler import TxtHandler

try:
    import psutil
except ImportError:
    psutil = None


def _thousands_br_int(n: int) -> str:
    s = f"{n}"
    parts = []
    i = len(s)
    while i > 0:
        parts.append(s[max(0, i - 3) : i])
        i -= 3
    return ".".join(reversed(parts))


def format_calculo_atualizado_br(value) -> str:
    v = float(value)
    for loc_name in ("pt_BR.UTF-8", "pt_BR"):
        try:
            locale.setlocale(locale.LC_ALL, loc_name)
            s = (
                locale.currency(v, grouping=True)
                .replace("R$", "")
                .replace(" ", "")
                .strip()
            )
            s = s.replace(",", ".", 1) if s.count(",") >= 1 else s
            if s and not s.startswith("="):
                return s
        except (ValueError, locale.Error):
            continue
    if v < 0:
        sign, v = "-", -v
    else:
        sign = ""
    vi = int(v + 1e-8)
    cents = int(round((v - vi) * 100))
    if cents >= 100:
        vi, cents = vi + 1, 0
    body = f"{_thousands_br_int(vi)},{cents:02d}"
    s = f"{sign}{body}".replace(",", ".", 1)
    return s


class CalculationAutomation:
    def __init__(self, main_dict, today: datetime) -> None:
        self.txt = TxtHandler()
        self._wb = None
        self._ws = None
        self.google_drive_link: str | None = None

        t0 = time.time()
        self.terminate_excel_process()
        print(f"\n\t[T] terminate_excel_process: {time.time() - t0:.2f}s")

        self.main_dict = main_dict
        self.tz = pytz.timezone("America/Sao_Paulo")
        self.today = today

        self.sheet_name = "(02.2) Mem Detalhada (DEPRE SP)"

        _calc_dir = os.path.dirname(os.path.abspath(__file__))
        self.main_file_path = os.path.join(
            _calc_dir,
            "Planilha de Cálculos V31 (Pós PEC 66) 01042026 (1).xlsm",
        )

        self.municipais_list = [
            "MUNICIPIO",
            "MUNICÍPIO",
            "MUNICIPAL",
            "MUNICIPIÁRIOS",
            "MUNICIPIARIOS",
            "MUNIC",
            "INSS",
        ]

        self.output_path = None
        self._last_calculo_value = None

    def _get_sheet(self, wb: openpyxl.workbook.workbook.Workbook):
        if self.sheet_name in wb.sheetnames:
            return wb[self.sheet_name]
        for name in wb.sheetnames:
            if self.sheet_name.strip() == name.strip():
                return wb[name]
        raise KeyError(
            f"Folha {self.sheet_name!r} não encontrada. Folhas: {wb.sheetnames}"
        )

    @staticmethod
    def _parse_number_from_cell(v) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v)
        if isinstance(v, str):
            s = v.strip()
            if not s or s.startswith("="):
                return None
            s = s.replace("R$", "").replace(" ", "").strip()
            s = s.replace(".", "").replace(",", ".") if re.search(r",\d{1,2}$", s) else s.replace(",", "")
            try:
                return float(s)
            except ValueError:
                return None
        return None

    def _read_o313_from_path(self, path: str) -> float | None:
        """Lê o total líquido com a mesma regra que ``memoria_calculo`` (incl. soma O309–O312)."""
        try:
            merged = load_merged_memoria_valores(
                path, self.sheet_name, print_ok=False
            )
            return total_liquido_arredondado(merged)
        except Exception as e:
            print(f"\n{Fore.YELLOW}[!] leitura memória / O313: {e}{Style.RESET_ALL}")
        return None

    def check_day(self):
        return self.today

    def send_to_drive(self, today: datetime):
        return

    def _upload_output_to_google_drive_if_configured(self) -> str | None:
        """
        Após ``SAVED`` em ``OUTPUT/``, envia o ``.xlsm`` para o Google Drive se
        ``GOOGLE_DRIVE_ENABLED=1`` e credenciais/pasta estiverem no ``.env``.
        Falha de rede não invalida o cálculo (só regista aviso).
        """
        p = self.output_path
        if not p or not os.path.isfile(p):
            return None
        try:
            from google_api.drive import upload_saved_spreadsheet

            return upload_saved_spreadsheet(p)
        except Exception as e:
            print(
                f"\n{Fore.YELLOW}[google_drive] Erro ao importar ou enviar (planilha local mantida): {e}"
                f"{Style.RESET_ALL}\n"
            )
            return None

    def remove_accentuation(self, input_str):
        nfkd_form = unicodedata.normalize("NFKD", input_str)
        return "".join(c for c in nfkd_form if not unicodedata.combining(c))

    def start_conn(self):
        try:
            h = DBHandler()
            return h if h.ok else None
        except Exception as e:
            print(f"\n[x] ERROR:\n\t{e}")
            return None

    def get_calculo_atualizado(self, id, db_handler):
        try:
            if self._last_calculo_value is not None:
                calculo_atualizado = self._last_calculo_value
                db_handler = self.start_conn()
                if not db_handler or not db_handler.ok:
                    self.update_error(id, None)
                    return
                calculo_atualizado_fmt = format_calculo_atualizado_br(
                    float(calculo_atualizado)
                )
                query = f"""
                    UPDATE precainfosnew
                    SET Calculo_Atualizado = '{calculo_atualizado_fmt}',
                        UPDATES_INDEX = UPDATES_INDEX + 1
                    WHERE id = {int(id)};
                """
                db_handler.cursor.execute(query)
                db_handler.config.commit()
                print(
                    f"\n\t{Fore.GREEN}Calculo Atualizado:"
                    f" {Fore.LIGHTGREEN_EX}R$ {calculo_atualizado_fmt}"
                    f"{Style.RESET_ALL}"
                )
                db_handler.cursor.close()
                db_handler.config.close()
                self.clean()
                return

            if not self.output_path or not os.path.isfile(self.output_path):
                self.update_error(id, None)
                return

            db_handler = self.start_conn()
            if not db_handler or not db_handler.ok:
                self.update_error(id, None)
                return
            try:
                n = self._read_o313_from_path(self.output_path)
                if n is None:
                    self.save_ERRORs(id)
                    print(
                        f"\n{Fore.YELLOW}[!] O313 / memória sem valor legível. "
                        f"Não foi gravado 'Calculo_Atualizado' (evita substituir por 0,00). "
                        f"Instale o LibreOffice no servidor, defina MEMORIA_LIBREOFFICE_BIN se "
                        f"o serviço correr como outro utilizador, e confirme o cache de fórmulas na planilha."
                        f"{Style.RESET_ALL}"
                    )
                else:
                    calculo_atualizado = format_calculo_atualizado_br(float(n))
                    query = f"""
                        UPDATE precainfosnew
                        SET Calculo_Atualizado = '{calculo_atualizado}',
                            UPDATES_INDEX = UPDATES_INDEX + 1
                        WHERE id = {int(id)};
                    """
                    db_handler.cursor.execute(query)
                    db_handler.config.commit()
                    print(
                        f"\n\t{Fore.GREEN}Calculo Atualizado:"
                        f" {Fore.LIGHTGREEN_EX}R$ {calculo_atualizado}"
                        f"{Style.RESET_ALL}"
                    )
            finally:
                try:
                    db_handler.cursor.close()
                    db_handler.config.close()
                except Exception:
                    pass
            self.clean()

        except Exception:
            self.update_error(id, None)
            traceback.print_exc()

    def save_ERRORs(self, id):
        current_datetime = datetime.now(self.tz).strftime("%d/%m/%Y %H:%M:%S")
        errors_path = os.path.join(os.path.dirname(__file__), "ERRORs")
        try:
            os.makedirs(errors_path, exist_ok=True)
            with open(
                os.path.join(errors_path, "ERRORs.txt"), "a", encoding="utf-8"
            ) as f:
                f.write(f"{current_datetime} - ID: {id}\n")
        except Exception as e:
            print(
                f"\n{Fore.YELLOW}[!] Não foi possível escrever ERRORs.txt em {errors_path!r}: {e}"
                f"{Style.RESET_ALL}"
            )

    def update_error(self, id, *_args):
        h = self.start_conn()
        if not h or not h.ok:
            print(f"\n{Fore.RED}[x] update_error: sem conexão MySQL{Style.RESET_ALL}")
            return
        query = f"""
            UPDATE precainfosnew
            SET Calculo_Atualizado = '0.00',
                UPDATES_INDEX = UPDATES_INDEX + 1
            WHERE id = {int(id)};
        """
        try:
            h.cursor.execute(query)
            h.config.commit()
            print(
                f"\n\t{Fore.BLUE}[ERRO] Calculo Atualizado:"
                f" {Fore.LIGHTGREEN_EX}R$ 0.00{Style.RESET_ALL}"
            )
        finally:
            try:
                h.end_conn()
            except Exception:
                pass

    def terminate_excel_process(self):
        if not psutil:
            return
        for proc in psutil.process_iter(["pid", "name"]):
            if proc.info["name"] and "excel" in proc.info["name"].lower():
                try:
                    psutil.Process(proc.info["pid"]).terminate()
                except Exception:
                    pass

    def clean(self):
        try:
            self.terminate_excel_process()
            _calc_dir = os.path.dirname(os.path.abspath(__file__))
            base_folder = _calc_dir
            output_folder = os.path.join(base_folder, "OUTPUT")
            bkp_folder = os.path.join(base_folder, "BKP")
            plans_folder = os.environ.get("PLANS_OUTPUT_DIR") or os.path.join(
                base_folder, "PLANS_ARCHIVED"
            )
            os.makedirs(plans_folder, exist_ok=True)
            os.makedirs(bkp_folder, exist_ok=True)

            if os.path.exists(bkp_folder):
                zip_files = [f for f in os.listdir(bkp_folder) if f.lower().endswith(".zip")]
                zip_files.sort(
                    key=lambda x: os.path.getmtime(os.path.join(bkp_folder, x))
                )
                while len(zip_files) > 10:
                    oldest_zip = zip_files.pop(0)
                    os.remove(os.path.join(bkp_folder, oldest_zip))
                    print(f"[INFO] Removed old zip file: {oldest_zip}")

            today_date = datetime.now().strftime("%d-%m-%Y")
            dated_plans_folder = os.path.join(plans_folder, today_date)
            os.makedirs(dated_plans_folder, exist_ok=True)

            def get_current_zip():
                zip_index = 1
                while True:
                    zip_name = os.path.join(bkp_folder, f"BKP_{zip_index}.zip")
                    if not os.path.exists(zip_name):
                        return zip_name
                    zip_size = os.path.getsize(zip_name) / (1024 * 1024)
                    if zip_size < 100:
                        return zip_name
                    zip_index += 1

            if os.path.exists(output_folder):
                for filename in os.listdir(output_folder):
                    file_path = os.path.join(output_folder, filename)
                    if os.path.isfile(file_path) and filename.lower().endswith(
                        (".xlsx", ".xlsm")
                    ):
                        zip_name = get_current_zip()
                        with zipfile.ZipFile(
                            zip_name, "a", zipfile.ZIP_DEFLATED
                        ) as zipf:
                            arcname = filename
                            if arcname not in zipf.namelist():
                                zipf.write(file_path, arcname=arcname)
                                zip_size = os.path.getsize(zip_name) / (1024 * 1024)
                                print(
                                    f"\n\t{Fore.LIGHTCYAN_EX}[INFO] Added: {filename} → "
                                    f"{os.path.basename(zip_name)} (Current size: {zip_size:.2f} MB){Style.RESET_ALL}\n"
                                )
                            else:
                                print(
                                    f"\n{Fore.YELLOW}[WARNING] Skipped duplicate in ZIP: {filename}{Style.RESET_ALL}\n"
                                )
                        dest_path = os.path.join(dated_plans_folder, filename)
                        shutil.move(file_path, dest_path)
                        print(
                            f"\n\t{Fore.LIGHTCYAN_EX}[INFO] Moved: {filename} → {dated_plans_folder}{Style.RESET_ALL}\n"
                        )
                    else:
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                            print(f"[INFO] Removed non-Excel file: {filename}")
        except Exception as e:
            print(f"\n{Fore.RED}[x] ERROR: clean()\n\t{e}{Style.RESET_ALL}")

    def save_workbook(self, proc_id):
        if self._wb is None or self._ws is None:
            return
        try:
            current_date = datetime.now().strftime("%d-%m-%Y")
            requerente = re.sub(
                r'[\/:*?"<>|()\\]',
                "",
                str(self.main_dict["Requerente"]).strip(),
            )
            requerente = re.sub(r"\s+", " ", requerente)
            # Por enquanto, geramos apenas .xlsx (melhor compatibilidade com Google Sheets).
            output_name = f"{proc_id}_{requerente}_{current_date}.xlsx"
            _calc_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(_calc_dir, "OUTPUT")
            os.makedirs(output_dir, exist_ok=True)
            self.output_path = os.path.join(output_dir, output_name)

            t_save = time.time()
            try:
                self._last_calculo_value = self._parse_number_from_cell(
                    self._ws["O313"].value
                )
            except Exception:
                self._last_calculo_value = None
            if self._last_calculo_value is None:
                try:
                    v = self._ws["O313"].value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        self._last_calculo_value = float(v)
                except Exception:
                    self._last_calculo_value = None

            self._wb.save(self.output_path)
            print(f"\n\t[T] workbook.save(): {time.time() - t_save:.2f}s")
            # Fechar antes do LibreOffice para evitar bloqueio do ficheiro .xlsm
            self._wb.close()
            self._wb = None
            self._ws = None

            # Saída é .xlsx, mas precisamos materializar fórmulas (O307–O313) antes de ler.
            # O LibreOffice headless abre e recalcula, exportando um .xlsx “com valores”.
            recalc_xlsx, recalc_tmpdir = libreoffice_recalc_export_xlsx(self.output_path)
            read_path = (recalc_xlsx or self.output_path).strip()
            if memoria_recalc_wanted() and not recalc_xlsx:
                raise RuntimeError(
                    "O LibreOffice não produziu ficheiro .xlsx recalculado — sem isso, "
                    "O307:O313 podem ficar vazios/errados e a BD ficaria mal preenchida. "
                    "Resolva: (1) instale, ex.: `sudo apt install -y libreoffice-calc`; "
                    "(2) teste `sudo -u www-data soffice --version`; "
                    "(3) no .env do projecto, se preciso, `MEMORIA_LIBREOFFICE_BIN=/usr/bin/soffice` "
                    "(ou o caminho de `command -v soffice`); (4) reinicie `atualizacao-calculo-api`. "
                    "Com `www-data`, defina também `MEMORIA_LIBREOFFICE_HOME` para um directório gravável "
                    "(ex. `/var/lib/lo-calc` com chown www-data) ou use o padrão "
                    "`calculation_automation/.lo_profile` com permissões correctas. "
                    "Apenas em teste, `MEMORIA_LIBREOFFICE=0` desactiva a conversão (não recomendado)."
                )
            try:
                merged = load_merged_memoria_valores(
                    read_path, self.sheet_name, print_ok=True
                )
                try:
                    sync_memoria_calculo_to_db(
                        self.main_dict,
                        self.output_path,
                        self.sheet_name,
                        read_memoria_path=read_path,
                        precomputed_merged=merged,
                        print_ok=True,
                    )
                except Exception as e:
                    print(
                        f"\n{Fore.YELLOW}[memoria_calculo] "
                        f"Falha ao gravar (registo já guardado em planilha): {e}{Style.RESET_ALL}\n"
                    )
                    traceback.print_exc()

                self._last_calculo_value = total_liquido_arredondado(merged)
            finally:
                if recalc_tmpdir and os.path.isdir(recalc_tmpdir):
                    try:
                        shutil.rmtree(recalc_tmpdir, ignore_errors=True)
                    except OSError:
                        pass
            if self._last_calculo_value is None:
                print(
                    f"\n{Fore.YELLOW}[!] O313 não pôde ser lido como número. "
                    f"Se LibreOffice não recalculou, instale/ajuste MEMORIA_LIBREOFFICE ou PATH. "
                    f"Verifique ERRORs.txt.{Style.RESET_ALL}\n"
                )
            print(
                f"\n\t{Fore.LIGHTBLACK_EX}SAVED {self.output_path}{Style.RESET_ALL}\n"
            )
            self.google_drive_link = self._upload_output_to_google_drive_if_configured()
        except Exception:
            traceback.print_exc()
            self._last_calculo_value = None
            try:
                if self._wb is not None:
                    self._wb.close()
            except Exception:
                pass
            self._wb = None
            self._ws = None
            raise

    def edit_cells(self):
        if not self.main_dict:
            return
        if not os.path.isfile(self.main_file_path):
            print(
                f"\n{Fore.RED}[x] Template não encontrado: {self.main_file_path}{Style.RESET_ALL}"
            )
            return

        t_edit_start = time.time()
        t_open = time.time()
        self._wb = load_workbook(
            # Template é .xlsm; manter VBA não é necessário quando a saída é .xlsx.
            self.main_file_path, keep_vba=False, data_only=False, read_only=False
        )
        self._ws = self._get_sheet(self._wb)
        print(f"\n\t[T] Template aberto (openpyxl): {time.time() - t_open:.2f}s")

        ws = self._ws
        entidade_devedora = self.remove_accentuation(
            str(self.main_dict["Entidade_Devedora"]).upper().strip().replace("  ", " ")
        )
        if any(term in entidade_devedora for term in ["INSS"]):
            entidade_devedora = "Federal-TJ"
        elif any(
            term in entidade_devedora
            for term in [
                "DE SAO PAULO",
                "DO ESTADO",
                "CBPM",
                "SPPREV",
                "ESTADUAL",
                "ESTADO",
            ]
        ):
            entidade_devedora = "Estadual/Municipal"
        elif any(
            term in entidade_devedora
            for term in ["PREFEITURA", "MUNICIPIO", "MUNICIPAL", "MUN."]
        ):
            entidade_devedora = "Estadual/Municipal"
        else:
            entidade_devedora = "Estadual/Municipal"

        reqte = str(self.main_dict["Requerente"]).strip()
        ws["B11"] = reqte
        ws["B21"] = reqte
        ws["B12"] = entidade_devedora
        ws["B14"] = self.main_dict["Processo"]
        ws["B15"] = self.main_dict["Oc"]
        ws["B16"] = self.main_dict["EP"]
        ws["B17"] = self.main_dict["Cumprimento"]
        ws["B18"] = self.main_dict["Incidente"]

        try:
            ws["F27"] = datetime.strptime(
                self.main_dict["Data_Base"], "%d/%m/%Y"
            )
        except Exception:
            pass
        try:
            ws["J27"] = datetime.strptime(
                self.main_dict["Data_Inscrição"], "%d/%m/%Y"
            )
        except Exception:
            pass

        previdencia_total = sum(
            [
                self.main_dict.get(k, 0)
                for k in [
                    "IPESP/IAMSP",
                    "SPPREV",
                    "IAMSPE",
                    "IPESP",
                    "ASSIT_MED_HOSPITAL",
                    "INST_PREV_CAIXA_BENEF",
                    "ASSIST_MED_CAIXA_BENEF",
                    "INST_PREV",
                ]
            ]
        )
        ws["B32"] = self.main_dict["Principal_Liquido"]
        ws["D32"] = self.main_dict["Juros_Moratorio"]
        ws["F32"] = self.main_dict["Despesas"]
        ws["I32"] = previdencia_total
        n_meses = int(self.main_dict.get("Numero_de_Meses", 0))
        ws["F310"] = n_meses if n_meses > 0 else "100.000.000"

        print(
            f"\n\t[T] edit_cells (preenchimento openpyxl): {time.time() - t_edit_start:.2f}s"
        )
        self.save_workbook(self.main_dict["id"])
