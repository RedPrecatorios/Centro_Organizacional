# -*- coding: utf-8 -*-
"""
Sincroniza ``memoria_calculo`` a partir do bloco **Memória de cálculo** da planilha
(células **O307–O313**), preferencialmente a partir de um **.xlsx** gerado com
**LibreOffice headless** (``--convert-to xlsx``) para materializar fórmulas; fallback:
leitura do ``.xlsm`` com ``data_only=True`` (cache vazio se nunca recalculado no Excel).

**Ambiente (opcional):** ``MEMORIA_LIBREOFFICE`` (``1``/``0``), ``MEMORIA_LIBREOFFICE_BIN``,
``MEMORIA_LIBREOFFICE_HOME`` (directório gravável para perfil dconf; obrigatório com ``User=www-data``:
por defeito ``calculation_automation/.lo_profile``), ``MEMORIA_LIBREOFFICE_TIMEOUT`` (s).

Mapeamento célula → coluna MySQL
--------------------------------

- **O307** → ``principal_bruto``
- **O308** → ``juros``
- **O309** → ``total_bruto``
- **O310** → ``desc_saude_prev`` (magnitude geralmente positiva na BD)
- **O311** → ``desc_ir`` (magnitude geralmente positiva na BD)
- **O312** → **valor monetário** da reserva de honorários → ``reserva_honorarios`` (pode ser negativo; ex. ``-R$ 63.346,23``)
- **O313** → ``total_liquido``; V31: ``=O309+O310+O311+O312``. O UPSERT grava a **soma** com
  **O309** (total bruto) obrigatório; parcelas O310–O312 em ``None`` contam como **0** na
  soma (p.ex. O311 sem cache). Leitura usa ``read_only=False`` para células unidas.

A coluna ``percentual_honorarios`` na BD (DEC ``30.00`` = 30 %), não a célula de valor; usa-se
``MEMORIA_HONORARIOS_PERCENT`` no ``.env`` (padrão **30.0**), alinhada ao rótulo “30 %” do Excel.

**Bases (``.env``):** ``_build_memoria_db_targets`` (MEMORIA_MYSQL_*, dúplica, legado).
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
import traceback
from typing import Any

import pymysql
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# 307–313: 7 células de dados (O); percentual de honorários = meta na BD, não célula O
O_COLUMN_TO_FIELD: tuple[tuple[str, str], ...] = (
    ("O307", "principal_bruto"),
    ("O308", "juros"),
    ("O309", "total_bruto"),
    ("O310", "desc_saude_prev"),
    ("O311", "desc_ir"),
    ("O312", "reserva_honorarios"),
    ("O313", "total_liquido"),
)


def _libreoffice_enabled() -> bool:
    v = (os.getenv("MEMORIA_LIBREOFFICE") or "1").strip().lower()
    return v not in ("0", "false", "no", "off", "disabled")


def memoria_recalc_wanted() -> bool:
    """
    Se True, espera-se conversão .xlsm → .xlsx com ``soffice`` para materializar fórmulas
    (O307:O313). Usado para falhar cedo com mensagem clara em vez de gravar zeros.
    """
    return _libreoffice_enabled()


def _find_libreoffice_cmd() -> str | None:
    for name in (
        (os.getenv("MEMORIA_LIBREOFFICE_BIN") or "").strip(),
        "libreoffice",
        "soffice",
    ):
        if not name:
            continue
        if os.path.isfile(name) and os.access(name, os.X_OK):
            return name
        path = shutil.which(name)
        if path:
            return path
    return None


def _libreoffice_recalc_timeout_sec() -> int:
    raw = (os.getenv("MEMORIA_LIBREOFFICE_TIMEOUT") or "300").strip()
    try:
        t = int(raw)
    except ValueError:
        t = 300
    return max(30, t)


def _libreoffice_profile_home() -> str:
    """
    HOME a passar ao ``soffice`` (dconf, perfil). ``www-data`` muitas vezes não escreve em
    ``/var/www/.cache`` — sem isto, soffice acaba com "User installation could not be completed".
    """
    raw = (os.getenv("MEMORIA_LIBREOFFICE_HOME") or "").strip()
    if raw:
        return os.path.abspath(os.path.expanduser(raw))
    here = os.path.dirname(os.path.abspath(__file__))
    project_lo = os.path.join(here, "..", "calculation_automation", ".lo_profile")
    return os.path.normpath(project_lo)


def _ensure_libreoffice_profile_dirs(home: str) -> None:
    for sub in (
        home,
        os.path.join(home, ".cache"),
        os.path.join(home, ".cache", "dconf"),
        os.path.join(home, ".config"),
    ):
        try:
            os.makedirs(sub, mode=0o700, exist_ok=True)
        except OSError:
            pass


def _libreoffice_subprocess_env() -> dict[str, str]:
    env = os.environ.copy()
    home = _libreoffice_profile_home()
    _ensure_libreoffice_profile_dirs(home)
    env["HOME"] = home
    env["XDG_CONFIG_HOME"] = os.path.join(home, ".config")
    env["XDG_CACHE_HOME"] = os.path.join(home, ".cache")
    # headless: evita procurar plugin X
    env.setdefault("SAL_USE_VCLPLUGIN", "svp")
    return env


def libreoffice_recalc_export_xlsx(xlsm_path: str) -> tuple[str | None, str | None]:
    """
    Converte ``.xlsm`` para ``.xlsx`` com LibreOffice em modo headless (recalcula ao abrir).

    Returns
    -------
    (xlsx_path, tempdir)
        Caminho do ``.xlsx`` a usar para leitura e directório a apagar com ``shutil.rmtree``,
        ou ``(None, None)`` se desactivado, binário inexistente ou falha.
    """
    if not _libreoffice_enabled():
        return None, None
    cmd = _find_libreoffice_cmd()
    if not cmd:
        print(
            "\n[memoria_calculo] LibreOffice não encontrado no PATH "
            "(instale o pacote `libreoffice` ou defina MEMORIA_LIBREOFFICE_BIN). "
            "A ler memória do .xlsm (cache de fórmulas pode estar vazio).\n"
        )
        return None, None
    if not xlsm_path or not os.path.isfile(xlsm_path):
        return None, None

    abspath = os.path.abspath(xlsm_path)
    out_base = os.path.splitext(os.path.basename(abspath))[0] + ".xlsx"
    tmpdir = tempfile.mkdtemp(prefix="lo_memoria_")
    timeout = _libreoffice_recalc_timeout_sec()
    lo_env = _libreoffice_subprocess_env()
    if not os.access(lo_env["HOME"], os.W_OK):
        print(
            f"\n[memoria_calculo] HOME para LibreOffice ({lo_env['HOME']!r}) sem escrita. "
            f"Crie a pasta, faça chown ao utilizador do serviço, ou defina "
            f"MEMORIA_LIBREOFFICE_HOME (ex. /var/lib/lo-calc) no .env.\n"
        )
    try:
        proc = subprocess.run(
            [
                cmd,
                "--headless",
                "--invisible",
                "--nologo",
                "--nofirststartwizard",
                "--nolockcheck",
                "--norestore",
                "--convert-to",
                "xlsx",
                "--outdir",
                tmpdir,
                abspath,
            ],
            capture_output=True,
            text=True,
            timeout=timeout,
            check=False,
            env=lo_env,
        )
    except subprocess.TimeoutExpired:
        print(
            f"\n[memoria_calculo] LibreOffice excedeu {timeout}s; "
            f"a apagar {tmpdir!r} e a usar o .xlsm.\n"
        )
        shutil.rmtree(tmpdir, ignore_errors=True)
        return None, None
    except OSError as e:
        print(f"\n[memoria_calculo] Erro a executar LibreOffice: {e}\n")
        shutil.rmtree(tmpdir, ignore_errors=True)
        return None, None

    out_path = os.path.join(tmpdir, out_base)
    if proc.returncode != 0 or not os.path.isfile(out_path):
        err = (proc.stderr or proc.stdout or "")[:2000]
        print(
            f"\n[memoria_calculo] LibreOffice falhou (rc={proc.returncode}): {err or '(sem saída)'}\n"
        )
        shutil.rmtree(tmpdir, ignore_errors=True)
        return None, None
    if os.path.getsize(out_path) < 32:
        shutil.rmtree(tmpdir, ignore_errors=True)
        return None, None

    print(
        f"\n[memoria_calculo] Recálculo via LibreOffice: {out_base!r} em {os.path.dirname(out_path)!r}.\n"
    )
    return out_path, tmpdir


def _clean_cell_to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value != value:  # NaN
            return None
        return float(value)
    s = str(value).strip()
    if not s or s in ("-", "—"):
        return None
    if s.startswith("="):
        return None
    if s.startswith("#") or "ERROR" in s.upper():
        return None
    neg_paren = s.startswith("(") and s.endswith(")")
    if neg_paren:
        s = s[1:-1].strip()
    s = s.replace("R$", "").replace(" ", "")
    if re.search(r",\d{1,2}$", s):
        s = s.replace(".", "").replace(",", ".", 1)
    else:
        s = s.replace(",", "")
    s = re.sub(r"[^0-9.+-eE]", "", s)
    if s in ("", "-", "+", "inf", "-inf"):
        return None
    if s in ("-inf", "inf", "inf-"):
        return None
    try:
        f = float(s)
        if neg_paren and f >= 0:
            f = -f
        return f
    except ValueError:
        return None


def _valor_para_coluna_db(field: str, v: float | None) -> float:
    if v is None:
        return 0.0
    if field in ("desc_saude_prev", "desc_ir"):
        return round(abs(v), 2)
    return round(float(v), 2)


def _get_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    for n in wb.sheetnames:
        if n.strip() == sheet_name.strip():
            return wb[n]
    raise KeyError(f"Folha {sheet_name!r} inexistente: {wb.sheetnames}")


def _celula_valor_se_unida(ws, coord: str) -> Any:
    """
    Em células unidas, só o canto sup.esq. tem valor; os restos são ``MergedCell`` sem
    ``.value`` útil. Exige planilha **não** ``read_only`` (``merged_cells`` inactivo nesse modo).
    """
    cell = ws[coord]
    if not isinstance(cell, MergedCell):
        return cell.value
    for mrange in ws.merged_cells.ranges:
        if cell.coordinate in mrange:
            return ws.cell(row=mrange.min_row, column=mrange.min_col).value
    return None


def read_memoria_valores_da_planilha(
    file_path: str, sheet_name: str, *, data_only: bool = True
) -> dict[str, float | None]:
    """
    Lê O307:O313 do ficheiro (após gravação, ``data_only=True`` usa cache de fórmulas
    se existir no ficheiro).
    """
    out: dict[str, float | None] = {}
    if not file_path or not os.path.isfile(file_path):
        return {f: None for _c, f in O_COLUMN_TO_FIELD}
    empty = {f: None for _c, f in O_COLUMN_TO_FIELD}
    try:
        is_xlsm = file_path.lower().endswith(".xlsm")
        # read_only=False: merged_cells activo; O310–O312 podem estar em ranges unidos
        # (Libre/Excel) e o valor cai só no canto.
        wb = load_workbook(
            file_path,
            data_only=data_only,
            read_only=False,
            keep_vba=is_xlsm,
        )
    except Exception as e:
        print(f"\n[memoria_calculo] Erro a abrir {file_path!r}: {e}\n")
        return empty
    try:
        ws = _get_sheet(wb, sheet_name)
        for coord, field in O_COLUMN_TO_FIELD:
            v = _celula_valor_se_unida(ws, coord)
            out[field] = _clean_cell_to_float(v)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return out


def _default_honorarios_percent() -> float:
    raw = (os.getenv("MEMORIA_HONORARIOS_PERCENT") or "30").strip()
    try:
        p = float(raw.replace(",", "."))
    except ValueError:
        p = 30.0
    if 0 < p <= 1.0:
        p = p * 100.0
    return min(100.0, max(0.0, round(p, 2)))


def _ensure_percent_0_100(v: float) -> float:
    if v <= 0:
        return 0.0
    if 0 < v <= 1.0:
        return round(v * 100, 2)
    if v > 100:
        return 100.0
    return round(v, 2)


def _memoria_mysql_connect_timeout() -> int:
    raw = (os.getenv("MEMORIA_MYSQL_CONNECT_TIMEOUT") or "120").strip()
    try:
        t = int(raw)
    except ValueError:
        t = 120
    return max(1, t)


def _build_memoria_db_targets() -> list[tuple[str, int, str, str, str, str]]:
    from db_handler.db_handler import _mysql_settings

    load_dotenv(override=True)
    out: list[tuple[str, int, str, str, str, str]] = []

    mdb = (os.getenv("MEMORIA_MYSQL_DATABASE") or os.getenv("MEMORIA_DB") or "").strip()
    if mdb:
        mhost = (os.getenv("MEMORIA_MYSQL_HOST") or "127.0.0.1").strip()
        try:
            mport = int(os.getenv("MEMORIA_MYSQL_PORT") or "3306")
        except ValueError:
            mport = 3306
        muser = (os.getenv("MEMORIA_MYSQL_USER") or "").strip()
        mpass = os.getenv("MEMORIA_MYSQL_PASSWORD") or ""
        if not muser:
            print(
                "\n[memoria_calculo] MEMORIA_MYSQL_DATABASE definido mas falta MEMORIA_MYSQL_USER.\n"
            )
        else:
            out.append((mhost, mport, mdb, muser, mpass, "MEMORIA_MYSQL (plataforma)"))

    ms = _mysql_settings()
    dup_def = (os.getenv("MEMORIA_MYSQL_DUP_TO_DEFAULT") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
        "sim",
    )

    def _append_default(label: str) -> None:
        if not ms.get("user") or not ms.get("database"):
            if dup_def and label.startswith("HOST"):
                print(
                    "\n[memoria_calculo] MEMORIA_MYSQL_DUP_TO_DEFAULT=1: defina no .env as "
                    "chaves de BD do Manager (DB, DB_USER/MYSQL_USER, PASS) como no db_handler.\n"
                )
            return
        out.append(
            (ms["host"], ms["port"], ms["database"], ms["user"], ms["password"], label)
        )

    if dup_def:
        _append_default("HOST/DB (Manager)")
    if not out:
        _append_default("HOST/DB (legado)")

    return out


def _merge_reads_prefer_data_only(
    from_file_do: dict[str, float | None],
    from_file_raw: dict[str, float | None],
) -> dict[str, float | None]:
    """Prefere cache ``data_only``; se None, tenta a mesma célula sem data_only (números directos)."""
    merged: dict[str, float | None] = {}
    for _c, f in O_COLUMN_TO_FIELD:
        a = from_file_do.get(f)
        b = from_file_raw.get(f) if from_file_raw else None
        merged[f] = a if a is not None else b
    return merged


def _apply_total_liquido_soma_O309_312(
    merged: dict[str, float | None], *, print_ok: bool
) -> None:
    """
    Fórmula V31: **O313** = ``=O309+O310+O311+O312``.

    O **O309** (``total_bruto``) tem de existir. Para **O310:O312**, se a leitura devolver
    ``None`` (cache, ``#N/A`` ou célula unida sem valor nesse canto), usa-se **0** no
    somatório (equivalente a “sem desconto” / parcela a zero), o que o Libre/openpyxl por
    vezes não materializa no mesmo sítio do Excel.
    """
    tb = merged.get("total_bruto")
    if tb is None:
        return
    ds = merged.get("desc_saude_prev")
    di = merged.get("desc_ir")
    rh = merged.get("reserva_honorarios")

    def _parcela_ou_zero(
        v: float | None, o_ref: str
    ) -> tuple[float, str | None]:
        if v is None:
            return 0.0, o_ref
        return float(v), None

    p310, a310 = _parcela_ou_zero(ds, "O310")
    p311, a311 = _parcela_ou_zero(di, "O311")
    p312, a312 = _parcela_ou_zero(rh, "O312")
    try:
        soma = float(tb) + p310 + p311 + p312
    except (TypeError, ValueError):
        return
    o313_antes = merged.get("total_liquido")
    merged["total_liquido"] = soma
    faltou = [x for x in (a310, a311, a312) if x is not None]
    if not print_ok:
        return
    if faltou:
        print(
            "\n\t[memoria_calculo] total_liquido: "
            f"{' '.join(faltou)} sem valor; tratados como 0 na soma (O313 = O309+O310+O311+O312). "
            f"Resultado: {soma:,.2f}.\n"
        )
    elif o313_antes is None:
        print(
            "\n\t[memoria_calculo] total_liquido: O313 vazio/erro; "
            f"usada soma O309+O310+O311+O312 = {soma:,.2f} (regra V31).\n"
        )
    else:
        try:
            a = float(o313_antes)
            if abs(a - soma) > 0.02:
                print(
                    "\n\t[memoria_calculo] total_liquido: O313 tinha "
                    f"{a:,.2f}; gravada a soma O309+…+O312 = {soma:,.2f}.\n"
                )
        except (TypeError, ValueError):
            print(
                f"\n\t[memoria_calculo] total_liquido: O313 ilegível; "
                f"usada soma O309+O310+O311+O312 = {soma:,.2f}.\n"
            )


def load_merged_memoria_valores(
    file_path: str, sheet_name: str, *, print_ok: bool = True
) -> dict[str, float | None]:
    """
    Lê O307:O313 (com merge data_only + raw) e define ``total_liquido`` pela soma
    O309+O310+O311+O312 (O310–O312 em ``None`` = 0 na soma se O309 existir).
    """
    r_do = read_memoria_valores_da_planilha(file_path, sheet_name, data_only=True)
    r_raw = read_memoria_valores_da_planilha(file_path, sheet_name, data_only=False)
    merged = _merge_reads_prefer_data_only(r_do, r_raw)
    _apply_total_liquido_soma_O309_312(merged, print_ok=print_ok)
    return merged


def total_liquido_arredondado(merged: dict[str, float | None]) -> float | None:
    """Valor de ``total_liquido`` após a mesma regra de arredondamento que o UPSERT."""
    v = merged.get("total_liquido")
    if v is None:
        return None
    return _valor_para_coluna_db("total_liquido", v)


def sync_memoria_calculo_to_db(
    main_dict: dict,
    output_path: str,
    sheet_name: str,
    *,
    read_memoria_path: str | None = None,
    precomputed_merged: dict[str, float | None] | None = None,
    print_ok: bool = True,
) -> bool:
    """
    UPSERT: identificação a partir de ``main_dict``; **valores monetários** só
    a partir de O307:O313 no ficheiro (por defeito ``output_path``; use
    ``read_memoria_path`` para um ``.xlsx`` recalculado pelo LibreOffice). Se
    ``precomputed_merged`` for passado, usa-o em vez de reler a planilha (ex.: o mesmo
    dicionário já usado para O313 / ``Calculo_Atualizado``).
    """
    targets = _build_memoria_db_targets()
    if not targets:
        print(
            "\n[memoria_calculo] Configure MEMORIA_MYSQL_* (banco da plataforma) ou "
            "HOST/DB/USER/PASS no .env.\n"
        )
        return False

    try:
        pid = int(main_dict.get("id") or 0)
    except (TypeError, ValueError):
        print("\n[memoria_calculo] id inválido; sincronização ignorada.\n")
        return False
    if pid <= 0:
        return False
    if not output_path or not os.path.isfile(output_path):
        print(
            f"\n[memoria_calculo] Ficheiro inexistente para leitura O307:O313: {output_path!r}\n"
        )
        return False

    read_from = (read_memoria_path or output_path).strip() or output_path
    if precomputed_merged is not None:
        merged = dict(precomputed_merged)
    else:
        if not read_from or not os.path.isfile(read_from):
            print(
                f"\n[memoria_calculo] Ficheiro de leitura inexistente: {read_from!r}\n"
            )
            return False
        merged = load_merged_memoria_valores(read_from, sheet_name, print_ok=print_ok)

    requerente = (str(main_dict.get("Requerente") or "")).strip()[:500]
    num_proc = (str(main_dict.get("Processo") or "")).strip()[:200]
    num_inc = (str(main_dict.get("Incidente") or "")).strip()[:200]

    if all(merged.get(f) is None for _c, f in O_COLUMN_TO_FIELD):
        print(
            "\n\t[memoria_calculo] Nenhum valor O307:O313 na planilha (cache vazio / LibreOffice inactivo). "
            "Não vou gravar memoria_calculo com zeros para não apagar dados anteriores. "
            "Instale o LibreOffice no servidor, confira MEMORIA_LIBREOFFICE e MEMORIA_LIBREOFFICE_BIN, "
            "e que o ficheiro .xlsx intermédio seja gerado.\n"
        )
        return False

    pb = _valor_para_coluna_db("principal_bruto", merged.get("principal_bruto"))
    juros = _valor_para_coluna_db("juros", merged.get("juros"))
    tb = _valor_para_coluna_db("total_bruto", merged.get("total_bruto"))
    ds = _valor_para_coluna_db("desc_saude_prev", merged.get("desc_saude_prev"))
    di = _valor_para_coluna_db("desc_ir", merged.get("desc_ir"))
    reserva = _valor_para_coluna_db("reserva_honorarios", merged.get("reserva_honorarios"))
    tliq = _valor_para_coluna_db("total_liquido", merged.get("total_liquido"))
    pct = _ensure_percent_0_100(_default_honorarios_percent())

    if print_ok and max(abs(pb), abs(juros), abs(tb), abs(tliq), abs(reserva)) < 1e-9:
        print(
            "\n\t[memoria_calculo] Aviso: O307:O313 vieram sem valores numéricos. "
            "Gravação na BD com zeros. Confirme LibreOffice, MEMORIA_LIBREOFFICE=1 e PATH; "
            "ou abra a planilha no Excel, grave, para preencher o cache de fórmulas.\n"
        )
    else:
        if print_ok:
            print(
                f"\n\t[memoria_calculo] A ler memória a partir de {os.path.basename(read_from)} "
                f"(O307:O313).\n"
            )

    sql = """
        INSERT INTO memoria_calculo (
            id_precainfosnew,
            requerente,
            numero_de_processo,
            numero_do_incidente,
            principal_bruto,
            juros,
            desc_saude_prev,
            desc_ir,
            percentual_honorarios,
            total_bruto,
            reserva_honorarios,
            total_liquido
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        ON DUPLICATE KEY UPDATE
            requerente = VALUES(requerente),
            numero_de_processo = VALUES(numero_de_processo),
            numero_do_incidente = VALUES(numero_do_incidente),
            principal_bruto = VALUES(principal_bruto),
            juros = VALUES(juros),
            desc_saude_prev = VALUES(desc_saude_prev),
            desc_ir = VALUES(desc_ir),
            percentual_honorarios = VALUES(percentual_honorarios),
            total_bruto = VALUES(total_bruto),
            reserva_honorarios = VALUES(reserva_honorarios),
            total_liquido = VALUES(total_liquido);
    """
    row = (
        pid,
        requerente or None,
        num_proc or None,
        num_inc or None,
        pb,
        juros,
        ds,
        di,
        pct,
        round(tb, 2),
        round(reserva, 2),
        round(tliq, 2),
    )

    ct = _memoria_mysql_connect_timeout()
    ok_any = False
    for host, port, database, user, password, label in targets:
        config = None
        try:
            config = pymysql.connect(
                host=host,
                port=port,
                database=database,
                user=user,
                password=password,
                charset="utf8mb4",
                connect_timeout=ct,
            )
            with config.cursor() as cur:
                cur.execute(sql, row)
            config.commit()
            ok_any = True
            if print_ok:
                print(
                    f"\n\t[memoria_calculo] [{label}] "
                    f"Registo salvo/actualizado: id_precainfosnew={pid} @ {database}\n"
                )
        except Exception as e:
            print(
                f"\n\t[memoria_calculo] Erro MySQL [{label}] "
                f"({host}:{port}/{database}): {e}\n"
            )
            traceback.print_exc()
        finally:
            if config:
                try:
                    config.close()
                except Exception:
                    pass

    return ok_any


def parse_memoria_n307_o313(sheet) -> dict[str, float]:  # noqa: ARG001
    """API antiga: leitura vem de ``sync_memoria_calculo_to_db`` + ficheiro gravado."""
    return {}


def parse_memoria_n306_o313(sheet) -> dict[str, float]:  # noqa: ARG001
    return parse_memoria_n307_o313(sheet)
